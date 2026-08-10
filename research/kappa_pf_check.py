"""Does power factor rank samples the same way zT does?

INTERNAL ONLY. Source data is research_1's (CS/CS-1/CS-3/CS-5), which is not
cleared for slides, papers or anything shown outside. No number produced here
may leave the room. The *method* conclusion is what we are after.

Why it matters: kappa comes from LFA, which is the slow expensive measurement.
S and rho come off the Linseis LSR in one sitting. If PF = S^2/rho ranks
candidates the same way zT does, a campaign can screen on PF and spend LFA time
only on the leaders -- more BO points for the same lab effort. If it does not,
that lever is dead for this material class.

Units, confirmed against the raw LSR header (Sensor Range lines):
    S   in microvolt/K      rho in micro-ohm*m      kappa in W/m/K
so PF in microwatt/m/K^2 is just S^2/rho, and zT = PF*1e-6 * T / kappa.
Everything below is recomputed from T, rho, kappa, S -- the sheet's own PF and
zT columns sit in different places on different sheets and are only used as a
cross-check that the columns were read correctly.
"""

from pathlib import Path

import numpy as np
import openpyxl
from scipy import stats

BOOK = Path(r"D:\Latos\data\demo_project\thermoelectric\zt_calc.xlsx")

# Column letters are identical for the four inputs on every sheet; only the
# derived columns wander, which is exactly why nothing derived is read.
COL_T, COL_RHO, COL_KAPPA, COL_S = 0, 1, 2, 3


def load() -> tuple[list[str], np.ndarray]:
    """Return sample labels and an array of shape (n_samples, n_temps, 4)."""
    wb = openpyxl.load_workbook(BOOK, data_only=True)
    labels, blocks = [], []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = row[COL_T : COL_S + 1]
            if any(v is None for v in vals):
                continue
            rows.append([float(v) for v in vals])
        labels.append(name.strip())
        blocks.append(np.asarray(rows))

    lengths = {b.shape[0] for b in blocks}
    if len(lengths) != 1:
        raise SystemExit(f"sheets have different row counts: {lengths}")
    return labels, np.stack(blocks)


def main() -> None:
    labels, data = load()
    temps = data[0, :, COL_T]
    if not np.allclose(data[:, :, COL_T], temps):
        raise SystemExit("sheets are not on a common temperature grid")

    rho, kappa, seebeck = data[..., COL_RHO], data[..., COL_KAPPA], data[..., COL_S]
    pf = seebeck**2 / rho                      # microwatt / m / K^2
    zt = pf * 1e-6 * temps[None, :] / kappa

    print(f"{len(labels)} samples x {len(temps)} temperatures "
          f"({temps.min():.0f}-{temps.max():.0f} K):  {', '.join(labels)}\n")

    # --- cross-check against the sheet's own numbers at the first row --------
    print("cross-check vs the workbook's own PF column (300 K row):")
    for i, lab in enumerate(labels):
        print(f"   {lab:<10} recomputed PF = {pf[i, 0]:8.2f} uW/m/K^2   "
              f"zT = {zt[i, 0]:.4f}")

    # --- the actual question -------------------------------------------------
    agree = 0
    rhos = []
    print("\n  T(K)   rank by PF (best first)      rank by zT (best first)     Spearman")
    print("  " + "-" * 76)
    for j, t in enumerate(temps):
        by_pf = [labels[k] for k in np.argsort(-pf[:, j])]
        by_zt = [labels[k] for k in np.argsort(-zt[:, j])]
        r = stats.spearmanr(pf[:, j], zt[:, j]).statistic
        rhos.append(r)
        same_top = by_pf[0] == by_zt[0]
        agree += same_top
        print(f"  {t:5.0f}   {' > '.join(by_pf):<28} {' > '.join(by_zt):<27} "
              f"{r:+.2f}{'   <-- same best' if same_top else ''}")

    rhos = np.asarray(rhos)
    print("\n" + "=" * 78)
    print(f"PF-best equals zT-best at {agree} of {len(temps)} temperatures "
          f"({100 * agree / len(temps):.0f}%)")
    print(f"Spearman rho across samples: median {np.median(rhos):+.2f}, "
          f"range {rhos.min():+.2f} to {rhos.max():+.2f}")

    # --- why: does kappa vary more than PF? ---------------------------------
    print("\nspread across samples (coefficient of variation, per temperature):")
    cv_pf = np.std(pf, axis=0) / np.mean(pf, axis=0)
    cv_k = np.std(kappa, axis=0) / np.mean(kappa, axis=0)
    print(f"   power factor  CV = {np.median(cv_pf):.2f} (median over T)")
    print(f"   kappa         CV = {np.median(cv_k):.2f}")
    print(f"   kappa range   = {kappa.min():.2f} to {kappa.max():.2f} W/m/K "
          f"({kappa.max() / kappa.min():.1f}x)")
    print("\nPF is a usable proxy only when kappa is near-constant across the "
          "design.\nIf kappa varies as much as or more than PF, it drives the "
          "ranking and PF cannot stand in for zT.")


if __name__ == "__main__":
    main()
