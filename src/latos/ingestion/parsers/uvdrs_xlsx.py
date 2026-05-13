"""UV-DRS parser for `.xlsx` workbooks.

File format
-----------
Excel (`.xlsx`). The user's UV-DRS workflow produces spreadsheets where:
- Column A: wavelength (nm)
- Column B: reflectance (% — scaled 0..100)
- Columns C+: derived quantities (Kubelka-Munk, Tauc plot inputs, etc.)

Multi-sheet workbooks are common: each sheet is a different sample
(e.g. "CS", "CS-1", "CS-3", ...). `parse_all()` returns one
`ParsedData` per non-empty sheet — the orchestrator creates one
`Measurement` per sheet, and the sheet name flows into
`metadata["sheet_name"]` so the sample-name heuristic can prefer it
over the folder-name fallback. `parse()` returns the first non-empty
sheet's data alone (for callers that only want a primary view).

Validation policy: see `xrd_rigaku_txt.py` — same contract.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, ClassVar

import numpy as np
import openpyxl

from latos.core.enums import Severity, Technique
from latos.core.models import ValidationIssue, utc_now
from latos.ingestion.base_parser import BaseParser
from latos.ingestion.parsed_data import ParsedData

__all__ = ["UvDrsXlsxParser"]

# Wavelength range UV-DRS instruments scan. Anything outside this range
# in column A means the file probably isn't UV-DRS.
_WAVELENGTH_NM_MIN = 100.0
_WAVELENGTH_NM_MAX = 2500.0

# How many cells to peek at when sniffing in `can_parse`. Larger → more
# robust against blank leading rows, smaller → faster.
_SNIFF_ROWS = 20

# Minimum UV-DRS-shaped rows in the sniff window required for a confident
# match. Three rows in the wavelength range with numeric col B is hard to
# fake by accident.
_MIN_UV_ROWS_FOR_MATCH = 3

# Each row needs at least this many columns to hold (wavelength, value).
_MIN_COLUMNS = 2


class UvDrsXlsxParser(BaseParser):
    """Parser for UV-DRS Excel `.xlsx` workbooks."""

    name: ClassVar[str] = "uvdrs-xlsx"
    # 1.0.2: parse_all returns one ParsedData per non-empty sheet; the
    # old single-sheet behaviour with a "skipped sheets" warning is
    # gone (every sheet is now its own measurement).
    version: ClassVar[str] = "1.0.2"
    technique: ClassVar[Technique] = Technique.UV_DRS
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx",)

    # ─── can_parse ───────────────────────────────────────────────────
    def can_parse(self, path: Path) -> float:
        """Confidence based on first-sheet first-column values being wavelengths.

        `.xlsx` is generic, so we sniff: open the first sheet, scan the
        first ~20 rows, and check that at least 3 rows have column A in
        the typical UV wavelength range (100..2500 nm) AND column B is
        numeric. That's specific enough to rule out random spreadsheets.
        """
        if not self._extension_matches(path):
            return 0.0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return 0.0
        try:
            sheet = wb[wb.sheetnames[0]]
            uv_rows = 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= _SNIFF_ROWS:
                    break
                if _looks_like_uvdrs_row(row):
                    uv_rows += 1
            return 1.0 if uv_rows >= _MIN_UV_ROWS_FOR_MATCH else 0.0
        finally:
            wb.close()

    # ─── parse / parse_all ───────────────────────────────────────────
    def parse(self, path: Path) -> ParsedData:
        """Parse the first non-empty sheet only.

        Kept as a primary-view shortcut for callers that don't care
        about multi-sheet workbooks. Most production callers should
        use `parse_all` instead — the orchestrator already does.
        """
        results = self.parse_all(path)
        if not results:
            return self._empty_result(
                [
                    ValidationIssue(
                        field="data",
                        severity=Severity.ERROR,
                        message="No non-empty sheets in workbook.",
                        detected_at=utc_now(),
                    ),
                ],
            )
        return results[0]

    def parse_all(self, path: Path) -> tuple[ParsedData, ...]:
        """Parse every non-empty sheet, one ParsedData per sheet.

        Empty / non-UV-DRS-shaped sheets are skipped silently — a
        workbook with a leading "Summary" or "Notes" sheet shouldn't
        blow up. If every sheet is empty, `parse()` surfaces a single
        error ParsedData (so the orchestrator records a failure rather
        than dropping the file silently).
        """
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return (
                self._empty_result(
                    [
                        ValidationIssue(
                            field="file",
                            severity=Severity.ERROR,
                            message=f"Could not open workbook: {exc}",
                            detected_at=utc_now(),
                        ),
                    ],
                ),
            )

        try:
            sheet_names = list(wb.sheetnames)
            results: list[ParsedData] = []
            for sheet_name in sheet_names:
                issues: list[ValidationIssue] = []
                sheet = wb[sheet_name]
                wavelength, reflectance = _read_sheet_first_two_columns(sheet, issues)
                if not wavelength:
                    # Empty / unparseable sheet — skip silently. We don't
                    # want one stray "Notes" sheet to produce a junk
                    # measurement.
                    continue
                results.append(
                    self._build_parsed_data(
                        sheet_name=sheet_name,
                        all_sheet_names=sheet_names,
                        wavelength=wavelength,
                        reflectance=reflectance,
                        issues=issues,
                    ),
                )
        finally:
            wb.close()

        return tuple(results)

    # ─── Internals ───────────────────────────────────────────────────
    def _build_parsed_data(
        self,
        *,
        sheet_name: str,
        all_sheet_names: list[str],
        wavelength: list[float],
        reflectance: list[float],
        issues: list[ValidationIssue],
    ) -> ParsedData:
        arrays = {
            "wavelength": np.asarray(wavelength, dtype=np.float64),
            "reflectance": np.asarray(reflectance, dtype=np.float64),
        }
        metadata: dict[str, Any] = {
            "sheet_name": sheet_name,
            "all_sheet_names": all_sheet_names,
            "n_points": len(wavelength),
            "wavelength_min_nm": min(wavelength),
            "wavelength_max_nm": max(wavelength),
        }
        return ParsedData(
            technique=self.technique,
            arrays=arrays,
            metadata=metadata,
            instrument="UV-DRS (xlsx export)",
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )

    def _empty_result(self, issues: list[ValidationIssue]) -> ParsedData:
        """Build a minimal ParsedData when parsing failed early."""
        return ParsedData(
            technique=self.technique,
            arrays={},
            metadata={},
            instrument="UV-DRS (xlsx export)",
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )


# ─── Module-level helpers ───────────────────────────────────────────
def _looks_like_uvdrs_row(row: tuple[Any, ...]) -> bool:
    """True if (col A, col B) look like (wavelength_nm, numeric)."""
    if len(row) < _MIN_COLUMNS:
        return False
    a, b = row[0], row[1]
    if not isinstance(a, int | float) or not isinstance(b, int | float):
        return False
    return _WAVELENGTH_NM_MIN <= float(a) <= _WAVELENGTH_NM_MAX


def _read_sheet_first_two_columns(
    sheet: Any,
    issues: list[ValidationIssue],
) -> tuple[list[float], list[float]]:
    """Pull (wavelength, reflectance) from columns A and B, skipping non-numeric rows.

    Empty leading rows are silently skipped. A row in the middle of the
    data section that fails to parse contributes one to the malformed
    count (collapsed into a single warning by the caller via `issues`).
    """
    wavelength: list[float] = []
    reflectance: list[float] = []
    malformed = 0

    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < _MIN_COLUMNS:
            continue
        a, b = row[0], row[1]
        if a is None or b is None:
            continue
        if not isinstance(a, int | float) or not isinstance(b, int | float):
            # Plausibly a header row; not malformed in itself, just skip.
            if wavelength:
                # Once we're inside the data block, a non-numeric row IS
                # malformed (we shouldn't see header text mid-stream).
                malformed += 1
            continue
        a_f = float(a)
        if not (_WAVELENGTH_NM_MIN <= a_f <= _WAVELENGTH_NM_MAX):
            # Out-of-range wavelength — probably some other tab content.
            if wavelength:
                malformed += 1
            continue
        wavelength.append(a_f)
        reflectance.append(float(b))

    if malformed > 0:
        issues.append(
            ValidationIssue(
                field="data",
                severity=Severity.WARNING,
                message=f"{malformed} row(s) inside the data block were not numeric; skipped.",
                detected_at=utc_now(),
            ),
        )

    return wavelength, reflectance
