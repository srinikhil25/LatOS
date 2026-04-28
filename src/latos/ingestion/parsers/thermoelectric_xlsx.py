"""Thermoelectric parser for `.xlsx` workbooks (zT-style summaries).

File format
-----------
Excel workbooks where each sheet is a different sample, and each row is
a temperature point. The header row contains column labels like
"Temperature (K)", "Resistivity (Ohm.m)", "Thermal conductivity
(Wm-1K-1)", "Seebeck Coefficient (microvolt/K)", "Powerfactor", "zT".

Column ORDER varies between sheets within the same workbook (older
exports added new derived columns rather than re-ordering), so we look
up every column by its header text rather than by position. Required
columns: Temperature + Resistivity + Thermal conductivity + Seebeck.
Optional: Power factor, zT.

Multi-sheet handling matches `UvDrsXlsxParser`: we parse only the first
sheet and emit a warning listing the other sheet names. Stage 2 will
expand a multi-sheet workbook into one measurement per sheet.

Validation policy: see `xrd_rigaku_txt.py` — same contract.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, ClassVar

import numpy as np
import openpyxl

from latos.core.enums import Severity, Technique
from latos.core.models import ValidationIssue, utc_now
from latos.ingestion.base_parser import BaseParser
from latos.ingestion.parsed_data import ParsedData

__all__ = ["ThermoelectricXlsxParser"]

# Sniff window — how many rows to peek at when sniffing the format.
_SNIFF_ROWS = 30

# Columns we look up by header substring. Each entry is
#   (output_array_name, [accepted header substring tokens])
# A header that contains ANY of the substrings (case-insensitive) is treated
# as that array. This lets the parser absorb minor wording drift between
# instrument exports without a brittle exact-match.
_COLUMN_MAP: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("temperature_k", ("temperature",)),
    ("resistivity_ohm_m", ("resistivity",)),
    ("thermal_conductivity", ("thermal conductivity",)),
    ("seebeck_uvk", ("seebeck",)),
    ("power_factor", ("powerfactor", "power factor")),
    ("zt", ("zt",)),
)

# Required columns — the first four. If any is missing, parsing returns
# an error.
_REQUIRED_COLUMNS = ("temperature_k", "resistivity_ohm_m", "thermal_conductivity", "seebeck_uvk")


class ThermoelectricXlsxParser(BaseParser):
    """Parser for thermoelectric-property `.xlsx` workbooks."""

    name: ClassVar[str] = "thermoelectric-xlsx"
    version: ClassVar[str] = "1.0.0"
    technique: ClassVar[Technique] = Technique.THERMOELECTRIC
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx",)

    # ─── can_parse ───────────────────────────────────────────────────
    def can_parse(self, path: Path) -> float:
        """Confidence based on the first sheet's header row containing TE keywords.

        At least 'Temperature' AND 'Seebeck' (or the case-insensitive
        substring `seebeck`) must appear in the same row. Both keywords
        together are very specific to thermoelectric exports.
        """
        if not self._extension_matches(path):
            return 0.0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return 0.0
        try:
            sheet = wb[wb.sheetnames[0]]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= _SNIFF_ROWS:
                    break
                row_text = " ".join(str(c).lower() for c in row if c is not None)
                if "temperature" in row_text and "seebeck" in row_text:
                    return 1.0
            return 0.0
        finally:
            wb.close()

    # ─── parse ───────────────────────────────────────────────────────
    def parse(self, path: Path) -> ParsedData:
        """Parse a thermoelectric `.xlsx` into a `ParsedData`."""
        issues: list[ValidationIssue] = []

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            issues.append(
                ValidationIssue(
                    field="file",
                    severity=Severity.ERROR,
                    message=f"Could not open workbook: {exc}",
                    detected_at=utc_now(),
                ),
            )
            return self._empty_result(issues)

        try:
            sheet_names = list(wb.sheetnames)
            sheet = wb[sheet_names[0]]
            arrays_dict, header_text = _read_te_sheet(sheet, issues)
        finally:
            wb.close()

        if len(sheet_names) > 1:
            others = sheet_names[1:]
            issues.append(
                ValidationIssue(
                    field="sheets",
                    severity=Severity.WARNING,
                    message=(
                        f"Workbook has {len(sheet_names)} sheets; only the first "
                        f"({sheet_names[0]!r}) was parsed. Skipped: {others}"
                    ),
                    detected_at=utc_now(),
                ),
            )

        # Required-column check — loud failure if essentials missing.
        missing = [k for k in _REQUIRED_COLUMNS if k not in arrays_dict]
        if missing:
            issues.append(
                ValidationIssue(
                    field="data",
                    severity=Severity.ERROR,
                    message=f"Required columns missing: {missing}",
                    detected_at=utc_now(),
                ),
            )

        arrays: dict[str, np.ndarray] = {
            name: np.asarray(values, dtype=np.float64) for name, values in arrays_dict.items()
        }

        metadata: dict[str, Any] = {
            "sheet_name": sheet_names[0],
            "all_sheet_names": sheet_names,
            "n_points": len(next(iter(arrays_dict.values()))) if arrays_dict else 0,
            "headers_found": header_text,
        }

        return ParsedData(
            technique=self.technique,
            arrays=arrays,
            metadata=metadata,
            instrument=None,
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )

    def _empty_result(self, issues: list[ValidationIssue]) -> ParsedData:
        """Build a minimal ParsedData when parsing failed early."""
        return ParsedData(
            technique=self.technique,
            arrays={},
            metadata={},
            instrument=None,
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )


# ─── Module-level helpers ───────────────────────────────────────────
def _read_te_sheet(
    sheet: Any,
    issues: list[ValidationIssue],
) -> tuple[dict[str, list[float]], list[str]]:
    """Locate the header row, map columns by name, return (arrays, header_text).

    Returns an empty dict and empty header list if no recognizable header
    row is present.
    """
    header_row_idx, col_map, header_text = _find_header_columns(sheet)
    if header_row_idx is None or not col_map:
        return {}, []

    # Read every data row below the header.
    arrays: dict[str, list[float]] = {key: [] for key in col_map}
    valid_n_rows = 0
    for r, row in enumerate(sheet.iter_rows(values_only=True)):
        if r <= header_row_idx:
            continue
        # Required: temperature must be numeric and present. Without it
        # we can't anchor the row to a temperature point.
        temp_idx = col_map.get("temperature_k")
        if temp_idx is None or temp_idx >= len(row):
            continue
        temp_val = row[temp_idx]
        if not isinstance(temp_val, int | float):
            continue

        # Collect all mapped columns. If ANY mapped column is missing/non-
        # numeric for this row, skip the entire row — same-length invariant.
        row_values: dict[str, float] = {}
        skip_row = False
        for key, idx in col_map.items():
            if idx >= len(row):
                skip_row = True
                break
            v = row[idx]
            if not isinstance(v, int | float):
                skip_row = True
                break
            row_values[key] = float(v)
        if skip_row:
            continue

        for key, val in row_values.items():
            arrays[key].append(val)
        valid_n_rows += 1

    if valid_n_rows == 0:
        issues.append(
            ValidationIssue(
                field="data",
                severity=Severity.ERROR,
                message="No data rows could be parsed (header found, but no numeric rows).",
                detected_at=utc_now(),
            ),
        )

    return arrays, header_text


def _find_header_columns(sheet: Any) -> tuple[int | None, dict[str, int], list[str]]:
    """Find the header row and map our known column names to their indices.

    Returns (header_row_index, {output_name: col_index}, raw_header_strings).
    """
    for r, row in enumerate(sheet.iter_rows(values_only=True)):
        # Header row: contains 'Temperature' AND 'Seebeck' (case-insensitive).
        row_text = " ".join(str(c).lower() for c in row if c is not None)
        if "temperature" not in row_text or "seebeck" not in row_text:
            continue

        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_lower = str(cell).lower()
            for output_name, tokens in _COLUMN_MAP:
                if output_name in col_map:
                    continue  # already filled
                if any(tok in cell_lower for tok in tokens):
                    col_map[output_name] = col_idx
                    break
        header_text = [str(c) for c in row if c is not None and str(c).strip()]
        return r, col_map, header_text

    return None, {}, []
