"""Tests for the one-command experimental cycle.

This is where the session's pieces meet: the workbook reader, the slope fit, the
per-point variance, the stopping verdict and the pre-registration. Most of what
follows checks the joins between them, because each was tested in isolation and
none of that proves they compose.

The pre-registration is the load-bearing part. It is the whole evidentiary basis
for the closed-loop claim, and it only counts if it was written before the
sample existed, so the tests treat "was it written, and does it contain the
prediction" as the thing that must not break.
"""

from __future__ import annotations

import json

import numpy as np
import openpyxl
import pytest

from latos.campaign_cycle import (
    CycleOutcome,
    SampleFit,
    aggregate_replicates,
    main,
    run_cycle,
)
from latos.ingestion.ite_workbook_template import (
    FIRST_DATA_ROW,
    HEADER_ROW,
    MEASUREMENTS_SHEET,
    SAMPLES_SHEET,
    write_template,
)
from latos.optimization import list_preregistrations, prereg_dir

DELTAS = (2.0, 5.0, 10.0)


def _campaign(tmp_path, samples, *, offset=0.35, noise=0.0, deltas=DELTAS):
    """Write a filled workbook. `samples` maps id -> (composition, true slope)."""
    path = write_template(tmp_path / "campaign.xlsx")
    wb = openpyxl.load_workbook(path)
    s, m = wb[SAMPLES_SHEET], wb[MEASUREMENTS_SHEET]
    sh = {c.value: c.column for c in s[HEADER_ROW] if c.value}
    mh = {c.value: c.column for c in m[HEADER_ROW] if c.value}
    rng = np.random.default_rng(0)

    row = FIRST_DATA_ROW
    for i, (sid, (x, slope)) in enumerate(samples.items()):
        r = FIRST_DATA_ROW + i
        s.cell(r, sh["sample_id"], sid)
        s.cell(r, sh["mass_IL_A_mg"], round(400.0 * x, 4))
        s.cell(r, sh["mass_IL_B_mg"], round(400.0 * (1.0 - x), 4))
        for j, dt in enumerate(deltas, start=1):
            m.cell(row, mh["meas_id"], f"{sid}-M{j}")
            m.cell(row, mh["sample_id"], sid)
            m.cell(row, mh["RH_percent"], 42.0)
            m.cell(row, mh["T_hot_C"], 25.0 + dt)
            m.cell(row, mh["T_cold_C"], 25.0)
            m.cell(row, mh["wait_time_s"], 1800)
            m.cell(row, mh["delta_V_mV"], slope * dt + offset + rng.normal(0.0, noise))
            m.cell(row, mh["electrode_material"], "gold")
            row += 1
    wb.save(path)
    return path


class TestTooLittleData:
    def test_two_samples_get_advice_rather_than_a_recommendation(self, tmp_path):
        """Three points is where a surrogate starts. Below it, say so.

        A recommendation from two samples would be the midpoint of the gap
        dressed up as a prediction.
        """
        path = _campaign(tmp_path, {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6)})
        outcome = run_cycle(path)
        assert outcome.result is None
        assert outcome.prereg_path is None
        assert any("both pure liquids and the midpoint" in m for m in outcome.messages)

    def test_a_sample_with_one_delta_t_contributes_nothing(self, tmp_path):
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.3)},
            deltas=(5.0,),
        )
        outcome = run_cycle(path)
        assert outcome.fits == ()
        assert any("contributed no value" in m for m in outcome.messages)

    def test_an_unreadable_workbook_is_reported_not_raised(self, tmp_path):
        broken = tmp_path / "broken.xlsx"
        broken.write_bytes(b"not a workbook")
        outcome = run_cycle(broken)
        assert outcome.fits == ()
        assert any("Could not read the workbook" in m for m in outcome.messages)


class TestAFullCycle:
    @pytest.fixture
    def outcome(self, tmp_path):
        path = _campaign(
            tmp_path,
            {
                "IL-001": (0.0, 1.1),
                "IL-002": (1.0, 1.6),
                "IL-003": (0.5, 2.35),
                "IL-004": (0.25, 1.8),
            },
            noise=0.04,
        )
        return run_cycle(path), path

    def test_every_sample_is_fitted(self, outcome):
        result, _ = outcome
        assert len(result.fits) == 4
        assert {f.sample_id for f in result.fits} == {"IL-001", "IL-002", "IL-003", "IL-004"}

    def test_the_slope_and_the_composition_are_both_recovered(self, outcome):
        result, _ = outcome
        by_id = {f.sample_id: f for f in result.fits}
        assert by_id["IL-003"].composition == pytest.approx(0.5)
        assert by_id["IL-003"].seebeck_mv_k == pytest.approx(2.35, abs=0.05)

    def test_the_electrode_offset_is_surfaced_per_sample(self, outcome):
        """The number a single-point measurement would have hidden."""
        result, _ = outcome
        assert all(f.offset_mv == pytest.approx(0.35, abs=0.1) for f in result.fits)

    def test_a_next_composition_is_recommended_inside_the_range(self, outcome):
        result, _ = outcome
        assert result.result is not None
        assert 0.0 <= result.result.recommendation.x <= 1.0

    def test_the_standard_errors_reach_the_optimizer(self, outcome):
        """The join that makes the reliability claim more than one bit per point."""
        result, _ = outcome
        assert result.result.config.point_noise_used is True

    def test_the_report_names_every_sample_and_the_verdict(self, outcome):
        result, _ = outcome
        text = result.report()
        for fit in result.fits:
            assert fit.sample_id in text
        assert "NEXT: mix at x =" in text
        assert result.result.stopping.action.upper() in text


class TestThePreRegistration:
    @pytest.fixture
    def written(self, tmp_path):
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)},
            noise=0.04,
        )
        return run_cycle(path), path

    def test_it_lands_where_the_validation_screen_reads(self, written):
        """Not merely "beside the workbook" — in the one directory readers use.

        This previously asserted `preregistrations/`, which is beside the
        workbook and which nothing reads. The test passed, the record was
        written, and the loop was open at its last joint.
        """
        outcome, path = written
        assert outcome.prereg_path is not None
        assert outcome.prereg_path.parent == prereg_dir(path.parent)

    def test_the_record_is_findable_by_the_reader_that_scores_it(self, written):
        """The join the path bug broke: writer and reader must agree.

        `list_preregistrations` is what the desktop app lists and what the
        outcome screen validates against. If a freeze made at the bench does
        not appear here, the pre-registration is evidence nobody can reach.
        """
        outcome, path = written
        found = list_preregistrations(path.parent)
        assert [e.path for e in found] == [str(outcome.prereg_path)]
        assert found[0].recommended_x == pytest.approx(outcome.result.recommendation.x)

    def test_it_records_the_prediction_and_its_interval(self, written):
        """Without these the record cannot be scored against the outcome."""
        outcome, _ = written
        record = json.loads(outcome.prereg_path.read_text(encoding="utf-8"))
        assert record["kind"] == "latos.bo.prereg"
        prediction = record["prediction_at_recommendation"]
        assert 0.0 <= prediction["x"] <= 1.0
        low, high = prediction["predictive_interval_95"]
        assert low <= prediction["predicted_mean"] <= high

    def test_it_records_whether_each_point_carried_its_own_variance(self, written):
        """Otherwise the record cannot explain the recommendation it freezes.

        A heteroscedastic fit weighs the same observations differently, so two
        runs could carry identical frozen configs and have reached different
        answers.
        """
        outcome, _ = written
        record = json.loads(outcome.prereg_path.read_text(encoding="utf-8"))
        assert record["frozen_config"]["point_noise_used"] is True

    def test_an_explicit_destination_is_honoured(self, tmp_path):
        path = _campaign(
            tmp_path, {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)}
        )
        target = tmp_path / "elsewhere"
        outcome = run_cycle(path, out_dir=target)
        assert outcome.prereg_path.parent == target

    def test_declining_to_freeze_writes_nothing_and_says_what_that_costs(self, tmp_path):
        """A preview must not leave a file that looks like evidence."""
        path = _campaign(
            tmp_path, {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)}
        )
        outcome = run_cycle(path, freeze_prereg=False)
        assert outcome.result is not None
        assert outcome.prereg_path is None
        assert not prereg_dir(path.parent).exists()
        assert any("cannot later be presented" in m for m in outcome.messages)


class TestSignDisagreement:
    def test_opposite_signs_change_the_target_and_the_report_says_so(self, tmp_path):
        """The case that would silently waste the whole budget.

        When the coefficient crosses zero the magnitude has an interior minimum,
        so optimising |S| walks to an endpoint the campaign already measured.
        """
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, -1.8), "IL-002": (1.0, 2.1), "IL-003": (0.5, 0.4)},
            noise=0.02,
        )
        outcome = run_cycle(path)
        assert any("crosses zero" in m for m in outcome.messages)
        assert any("crossing" in m for m in outcome.messages)

    def test_one_sided_data_says_nothing_about_a_crossing(self, tmp_path):
        path = _campaign(
            tmp_path, {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)}
        )
        outcome = run_cycle(path)
        assert not any("crosses zero" in m for m in outcome.messages)


class TestDegenerateUncertainty:
    def test_perfectly_linear_data_is_refused_as_a_noise_estimate(self, tmp_path):
        """Noiseless input is not precision, it is the absence of an estimate.

        Passing it through set the convergence floor and the epsilon tolerance to
        roughly zero, and the engine then reported being within 1e-16 of the
        optimum — a statement about nothing. Found by running the command on
        synthetic data, not by reasoning about it.
        """
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)},
            noise=0.0,
        )
        outcome = run_cycle(path)
        assert any("no measurement-noise estimate" in m for m in outcome.messages)
        assert outcome.result.config.point_noise_used is False
        assert outcome.result.epsilon > 1e-6


class TestTheCommandLine:
    def test_it_prints_a_report_and_succeeds(self, tmp_path, capsys):
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)},
            noise=0.04,
        )
        assert main([str(path)]) == 0
        assert "NEXT: mix at x =" in capsys.readouterr().out

    def test_dry_run_leaves_no_file_behind(self, tmp_path, capsys):
        path = _campaign(
            tmp_path,
            {"IL-001": (0.0, 1.1), "IL-002": (1.0, 1.6), "IL-003": (0.5, 2.35)},
            noise=0.04,
        )
        assert main([str(path), "--dry-run"]) == 0
        capsys.readouterr()
        assert not prereg_dir(path.parent).exists()

    def test_a_missing_file_fails_without_a_traceback(self, tmp_path, capsys):
        assert main([str(tmp_path / "nope.xlsx")]) == 2
        assert "No such workbook" in capsys.readouterr().out

    def test_an_unusable_workbook_exits_nonzero(self, tmp_path, capsys):
        path = _campaign(tmp_path, {"IL-001": (0.0, 1.1)}, deltas=(5.0,))
        assert main([str(path)]) == 1
        capsys.readouterr()


class TestReportWithoutAResult:
    def test_a_report_is_still_printable_when_nothing_could_be_fitted(self):
        """The command must say something useful on its worst day."""
        outcome = CycleOutcome((), None, None, ("nothing to do",))
        text = outcome.report()
        assert "(none usable)" in text
        assert "nothing to do" in text


class TestReplicateAggregation:
    """Independent specimens at one composition are one observation, not several.

    The point of aggregating them is not tidiness. A slope-fit standard error
    describes how well a line was drawn through one specimen; the scatter
    between specimens made the same way describes whether making it again gives
    the same answer, and that is the quantity the surrogate needs. It is also
    the larger of the two, because it carries the weighing, the mixing and the
    mounting as well as the voltmeter.
    """

    @staticmethod
    def _fit(sample_id, composition, slope, stderr=0.002):
        return SampleFit(
            sample_id=sample_id,
            composition=composition,
            seebeck_mv_k=slope,
            stderr_mv_k=stderr,
            offset_mv=0.0,
            n_points=5,
            notes=(),
        )

    def test_replicates_collapse_to_one_point_carrying_their_mean(self):
        fits = (
            self._fit("A1", 0.0, 1.00),
            self._fit("A2", 0.0, 1.20),
            self._fit("B", 0.5, 2.00),
            self._fit("C", 1.0, 1.40),
        )
        points, _ = aggregate_replicates(fits)
        assert len(points) == 3
        first = points[0]
        assert first.n_replicates == 2
        assert first.value == pytest.approx(1.10)
        assert first.sample_ids == ("A1", "A2")

    def test_the_uncertainty_comes_from_replicate_scatter_not_the_fit(self):
        """The fit errors are tiny and the specimens disagree. Believe the specimens."""
        fits = (
            self._fit("A1", 0.0, 1.00, stderr=1e-4),
            self._fit("A2", 0.0, 1.20, stderr=1e-4),
            self._fit("B", 0.5, 2.00, stderr=1e-4),
            self._fit("C", 1.0, 1.40, stderr=1e-4),
        )
        points, _ = aggregate_replicates(fits)
        replicated = points[0]
        assert replicated.sigma_source == "replicates"
        # sd of (1.00, 1.20) is 0.1414; the mean of two carries sd/sqrt(2).
        assert replicated.sigma == pytest.approx(0.1414 / np.sqrt(2), rel=1e-3)
        assert replicated.sigma > 1e-4 * 100  # nothing like the fit error

    def test_an_unreplicated_point_inherits_the_pooled_spread(self):
        """One specimen tells you nothing about its own reproducibility.

        The right uncertainty for it is the spread measured everywhere else,
        not its own fit error, which describes a different thing entirely.
        """
        fits = (
            self._fit("A1", 0.0, 1.00, stderr=1e-4),
            self._fit("A2", 0.0, 1.20, stderr=1e-4),
            self._fit("B", 0.5, 2.00, stderr=1e-4),
        )
        points, _ = aggregate_replicates(fits)
        lone = next(p for p in points if p.n_replicates == 1)
        assert lone.sigma_source == "pooled"
        assert lone.sigma == pytest.approx(0.1414, rel=1e-3)

    def test_without_replicates_the_previous_behaviour_is_unchanged(self):
        fits = (
            self._fit("A", 0.0, 1.00, stderr=0.011),
            self._fit("B", 0.5, 2.00, stderr=0.022),
            self._fit("C", 1.0, 1.40, stderr=0.033),
        )
        points, notes = aggregate_replicates(fits)
        assert [p.sigma for p in points] == [0.011, 0.022, 0.033]
        assert all(p.sigma_source == "fit" for p in points)
        assert notes == []

    def test_compositions_within_weighing_tolerance_are_one_condition(self):
        """30.00 % and 30.02 % is one condition attempted twice."""
        fits = (
            self._fit("A1", 0.300, 1.00),
            self._fit("A2", 0.3002, 1.10),
            self._fit("B", 0.6, 2.00),
            self._fit("C", 0.9, 1.40),
        )
        points, _ = aggregate_replicates(fits)
        assert len(points) == 3
        assert points[0].n_replicates == 2

    def test_a_sign_disagreement_between_replicates_is_flagged_not_averaged(self):
        """Two specimens of opposite carrier type is a fault, not scatter."""
        fits = (
            self._fit("A1", 0.0, +2.0),
            self._fit("A2", 0.0, -2.0),
            self._fit("B", 0.5, 1.0),
            self._fit("C", 1.0, 1.4),
        )
        points, _ = aggregate_replicates(fits)
        assert any("SIGN" in note for note in points[0].notes)

    def test_identical_replicates_are_refused_as_a_noise_estimate(self):
        """Perfect agreement between specimens means computed, not measured."""
        fits = (
            self._fit("A1", 0.0, 1.0),
            self._fit("A2", 0.0, 1.0),
            self._fit("B", 0.5, 2.0),
            self._fit("C", 1.0, 1.4),
        )
        _, notes = aggregate_replicates(fits)
        assert any("computed" in note for note in notes)

    def test_few_degrees_of_freedom_are_declared_rather_than_hidden(self):
        fits = (
            self._fit("A1", 0.0, 1.0),
            self._fit("A2", 0.0, 1.2),
            self._fit("B", 0.5, 2.0),
            self._fit("C", 1.0, 1.4),
        )
        _, notes = aggregate_replicates(fits)
        assert any("itself uncertain" in note for note in notes)

    def test_replicates_of_one_composition_cannot_fit_a_surrogate(self, tmp_path):
        """Nine samples at three compositions is three points, not nine.

        The gate is on distinct conditions, because replicating one condition
        nine times teaches the surrogate nothing about the other end of the range.
        """
        path = _campaign(
            tmp_path,
            {"IL-001": (0.5, 1.1), "IL-002": (0.5, 1.2), "IL-003": (0.5, 1.15)},
        )
        outcome = run_cycle(path)
        assert outcome.result is None
        assert len(outcome.points) == 1
        assert any("DISTINCT compositions" in m for m in outcome.messages)

    def test_a_replicated_campaign_runs_end_to_end(self, tmp_path):
        path = _campaign(
            tmp_path,
            {
                "IL-001": (0.0, 1.10),
                "IL-002": (0.0, 1.22),
                "IL-003": (0.5, 2.30),
                "IL-004": (0.5, 2.18),
                "IL-005": (1.0, 1.60),
            },
        )
        outcome = run_cycle(path)
        assert outcome.result is not None
        assert len(outcome.fits) == 5
        assert len(outcome.points) == 3
        assert any(p.sigma_source == "replicates" for p in outcome.points)
        assert any("pooled standard deviation" in m for m in outcome.messages)
        assert "Design points given to the surrogate" in outcome.report()
