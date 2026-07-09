"""Tests for `latos.ingestion.parsers.lfa_xlsx.LfaXlsxParser`."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from latos.core.enums import Technique
from latos.ingestion.parsers.lfa_xlsx import LfaXlsxParser

_HEADER = [
    "#Temperature/K",
    "#Model",
    "#Diffusivity/(mm^2/s)",
    "#Conductivity/(W/(m*K))",
    "#Cp-Calc/(J/(g*K))",
]
_ROWS = [
    (300, "Standard + p.c.(l)", 2.574, 5.1455, 0.3501),
    (325, "Standard + p.c.(l)", 2.458, 5.1206, 0.3648),
    (350, "Standard + p.c.(l)", 2.367, 4.9464, 0.3660),
]


def _write_lfa(path: Path, *, blank_leading: int = 3) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(blank_leading):
        ws.append([])
    ws.append(_HEADER)
    for r in _ROWS:
        ws.append(list(r))
    wb.save(path)
    return path


@pytest.fixture()
def lfa_file(tmp_path: Path) -> Path:
    return _write_lfa(tmp_path / "CS LFA.xlsx")


class TestCanParse:
    def test_matches_lfa_workbook(self, lfa_file: Path):
        assert LfaXlsxParser().can_parse(lfa_file) == 1.0

    def test_rejects_non_lfa_xlsx(self, tmp_path: Path):
        wb = openpyxl.Workbook()
        wb.active.append(["wavelength", "reflectance"])
        wb.active.append([400, 12.3])
        p = tmp_path / "other.xlsx"
        wb.save(p)
        assert LfaXlsxParser().can_parse(p) == 0.0

    def test_rejects_wrong_extension(self, tmp_path: Path):
        p = tmp_path / "CS LFA.txt"
        p.write_text("x")
        assert LfaXlsxParser().can_parse(p) == 0.0


class TestParse:
    def test_extracts_conductivity_and_temperature(self, lfa_file: Path):
        d = LfaXlsxParser().parse(lfa_file)
        assert d.technique is Technique.THERMOELECTRIC
        assert list(d.arrays["temperature_k"]) == [300, 325, 350]
        assert d.arrays["thermal_conductivity"][0] == pytest.approx(5.1455)
        assert d.arrays["diffusivity_mm2_s"][0] == pytest.approx(2.574)
        assert d.metadata["measurement_kind"] == "lfa"
        assert d.metadata["n_points"] == 3

    def test_sample_name_stripped_from_filename(self, lfa_file: Path):
        d = LfaXlsxParser().parse(lfa_file)
        assert d.metadata["sample_name"] == "CS"

    def test_sample_name_keeps_doping_label(self, tmp_path: Path):
        f = _write_lfa(tmp_path / "CS-CBI-1 LFA.xlsx")
        d = LfaXlsxParser().parse(f)
        assert d.metadata["sample_name"] == "CS-CBI-1"

    def test_empty_below_header_errors(self, tmp_path: Path):
        wb = openpyxl.Workbook()
        wb.active.append(_HEADER)  # header only, no data
        p = tmp_path / "CS LFA.xlsx"
        wb.save(p)
        d = LfaXlsxParser().parse(p)
        assert d.arrays == {}
        assert any(i.field == "data" for i in d.issues)
