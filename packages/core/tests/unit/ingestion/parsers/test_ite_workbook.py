"""Tests for the ionic-thermoelectric recording workbook reader.

The workbook is the contract between the bench and the code, so most of what
follows is about what happens when a human fills it in imperfectly: a blank
humidity cell, a measurement row pointing at a sample that was never prepared,
a series too short to fit a slope. None of those may be silently dropped or
silently defaulted, because the whole reason for the Tier-1 fields is that they
cannot be reconstructed after the run.
"""

from __future__ import annotations

from datetime import UTC, datetime

import numpy as np
import openpyxl
import pytest

from latos.core.enums import Severity, Technique
from latos.ingestion.ite_workbook_template import (
    MEASUREMENT_COLUMNS,
    SAMPLE_COLUMNS,
    write_template,
)
from latos.ingestion.parsers.ite_workbook import IteWorkbookParser, _first_datetime

SAMPLE_COLS = [
    "sample_id",
    "IL_A_name",
    "IL_B_name",
    "mass_IL_A_mg",
    "mass_IL_B_mg",
    "fabric_type",
    "predicted_S_mV_K",
    "predicted_S_sigma",
    "notes",
]
MEAS_COLS = [
    "meas_id",
    "sample_id",
    "datetime_start",
    "RH_percent",
    "T_ambient_C",
    "T_hot_C",
    "T_cold_C",
    "wait_time_s",
    "steady_state_reached",
    "delta_V_mV",
    "electrode_material",
    "electrode_spacing_mm",
    "voltmeter_model",
    "raw_trace_file",
]


def _write(tmp_path, samples, measurements, *, sheets=("samples", "measurements")):
    """Build a workbook with the template's layout: header row 2, data row 4."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cols, rows in (
        (sheets[0], SAMPLE_COLS, samples),
        (sheets[1], MEAS_COLS, measurements),
    ):
        ws = wb.create_sheet(name)
        ws.append(["one row per thing"])  # row 1: title
        ws.append(cols)  # row 2: header
        ws.append([""] * len(cols))  # row 3: units
        for row in rows:
            ws.append([row.get(c) for c in cols])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def _sample(sid="IL-001", **over):
    base = {
        "sample_id": sid,
        "IL_A_name": "[EMIM][TFSI]",
        "IL_B_name": "[BMIM][Cl]",
        "mass_IL_A_mg": 300.0,
        "mass_IL_B_mg": 100.0,
        "fabric_type": "cotton",
        "predicted_S_mV_K": 2.4,
        "predicted_S_sigma": 0.3,
    }
    return base | over


def _meas(sid="IL-001", mid="M-1", dt=5.0, dv=12.0, **over):
    base = {
        "meas_id": mid,
        "sample_id": sid,
        "RH_percent": 42.0,
        "T_ambient_C": 23.0,
        "T_hot_C": 25.0 + dt,
        "T_cold_C": 25.0,
        "wait_time_s": 1800,
        "steady_state_reached": "yes",
        "delta_V_mV": dv,
        "electrode_material": "gold",
        "electrode_spacing_mm": 2.0,
        "voltmeter_model": "Keithley 6517B",
        "raw_trace_file": f"traces/{mid}.csv",
    }
    return base | over


def _series(sid="IL-001", slope=2.4, offset=0.0):
    return [
        _meas(sid, f"M-{i}", dt=dt, dv=slope * dt + offset)
        for i, dt in enumerate((2.0, 5.0, 10.0), start=1)
    ]


class TestRecognition:
    def test_a_filled_template_is_recognised_unambiguously(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series())
        assert IteWorkbookParser().can_parse(path) == 1.0

    def test_an_unrelated_workbook_is_declined(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.append(["temperature", "resistivity"])
        path = tmp_path / "other.xlsx"
        wb.save(path)
        assert IteWorkbookParser().can_parse(path) == 0.0

    def test_sheet_names_alone_are_not_enough(self, tmp_path):
        """Plenty of workbooks have a sheet called "samples"."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name in ("samples", "measurements"):
            wb.create_sheet(name).append(["something", "else"])
        path = tmp_path / "decoy.xlsx"
        wb.save(path)
        assert IteWorkbookParser().can_parse(path) == 0.0


class TestOneMeasurementPerSample:
    def test_each_sample_row_becomes_its_own_parsed_result(self, tmp_path):
        path = _write(
            tmp_path,
            [_sample("IL-001"), _sample("IL-002", mass_IL_A_mg=100.0, mass_IL_B_mg=300.0)],
            _series("IL-001") + _series("IL-002"),
        )
        results = IteWorkbookParser().parse_all(path)
        assert len(results) == 2
        assert [r.metadata["sample_id"] for r in results] == ["IL-001", "IL-002"]

    def test_the_delta_series_is_carried_as_arrays(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series(slope=2.4, offset=1.0))
        (result,) = IteWorkbookParser().parse_all(path)
        assert np.allclose(result.arrays["delta_t_k"], [2.0, 5.0, 10.0])
        assert np.allclose(result.arrays["delta_v_mv"], [5.8, 13.0, 25.0])

    def test_the_arrays_feed_the_slope_analyzer_directly(self, tmp_path):
        """The join between the two modules, checked rather than assumed."""
        from latos.analysis.thermovoltage.slope import fit_seebeck_slope

        path = _write(tmp_path, [_sample()], _series(slope=2.4, offset=1.0))
        (result,) = IteWorkbookParser().parse_all(path)
        fit = fit_seebeck_slope(result.arrays["delta_t_k"], result.arrays["delta_v_mv"])
        assert fit.slope == pytest.approx(2.4)
        assert fit.intercept == pytest.approx(1.0)

    def test_technique_and_parser_identity(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.technique is Technique.THERMOELECTRIC
        assert result.parser_name == "ite-workbook"


class TestDerivedComposition:
    def test_mass_fraction_is_recomputed_from_the_weighed_masses(self, tmp_path):
        """Recomputed, not read from the sheet's own formula column.

        A stale formula or a hand-typed value must not be able to disagree with
        the masses sitting beside it.
        """
        path = _write(tmp_path, [_sample(mass_IL_A_mg=300.0, mass_IL_B_mg=100.0)], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.metadata["mass_fraction_x"] == pytest.approx(0.75)
        assert result.features["mass_fraction_x"] == pytest.approx(0.75)

    def test_the_preregistered_prediction_is_surfaced(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.features["predicted_S_mV_K"] == pytest.approx(2.4)
        assert result.features["predicted_S_sigma"] == pytest.approx(0.3)

    def test_missing_masses_are_reported_not_guessed(self, tmp_path):
        path = _write(tmp_path, [_sample(mass_IL_B_mg=None)], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert "mass_fraction_x" not in result.metadata
        assert any(i.field == "mass_IL_B_mg" for i in result.issues)


class TestTier1Enforcement:
    """The fields that cannot be recovered once the bench is tidied."""

    @pytest.mark.parametrize(
        "field", ["RH_percent", "wait_time_s", "electrode_material", "T_hot_C"]
    )
    def test_a_blank_tier1_cell_is_reported_against_its_own_field(self, tmp_path, field):
        rows = _series()
        rows[1][field] = None
        path = _write(tmp_path, [_sample()], rows)
        (result,) = IteWorkbookParser().parse_all(path)
        matching = [i for i in result.issues if i.field == field]
        assert matching, f"{field} went unreported"
        assert "M-2" in matching[0].message

    def test_a_missing_humidity_does_not_discard_the_point(self, tmp_path):
        """Humidity is needed to interpret the value, not to compute it.

        Dropping the row would lose a perfectly good (ΔT, ΔV) pair; keeping it
        without the flag would hide that the series cannot be checked for a
        humidity confound later.
        """
        rows = _series()
        rows[1]["RH_percent"] = None
        path = _write(tmp_path, [_sample()], rows)
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.arrays["delta_t_k"].size == 3
        assert any(i.field == "RH_percent" for i in result.issues)

    def test_a_row_without_a_voltage_cannot_contribute_a_point(self, tmp_path):
        rows = _series()
        rows[1]["delta_V_mV"] = None
        path = _write(tmp_path, [_sample()], rows)
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.arrays["delta_t_k"].size == 2
        assert any(i.field == "delta_V_mV" for i in result.issues)


class TestSeriesTooShort:
    def test_a_single_point_yields_no_arrays_but_is_still_emitted(self, tmp_path):
        """A sample that vanishes is indistinguishable from one never prepared."""
        path = _write(tmp_path, [_sample()], [_meas()])
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.arrays == {}
        assert result.metadata["sample_id"] == "IL-001"
        assert any("fitted slope" in i.message for i in result.issues)

    def test_a_sample_with_no_measurements_at_all_is_still_emitted(self, tmp_path):
        path = _write(tmp_path, [_sample("IL-001"), _sample("IL-002")], _series("IL-001"))
        results = IteWorkbookParser().parse_all(path)
        assert len(results) == 2
        assert results[1].arrays == {}


class TestOrphanRows:
    def test_measurements_pointing_at_an_unknown_sample_are_reported(self, tmp_path):
        """A typo in a sample id would otherwise silently lose real data."""
        path = _write(tmp_path, [_sample("IL-001")], [*_series("IL-001"), _meas("IL-999")])
        (result,) = IteWorkbookParser().parse_all(path)
        orphan_issues = [i for i in result.issues if "IL-999" in i.message]
        assert orphan_issues
        assert orphan_issues[0].severity is Severity.WARNING


class TestPerPointMetadata:
    def test_humidity_is_kept_per_point_rather_than_averaged(self, tmp_path):
        """Drift across a series is the confound; a mean would erase it."""
        rows = _series()
        for row, rh in zip(rows, (30.0, 45.0, 60.0), strict=True):
            row["RH_percent"] = rh
        path = _write(tmp_path, [_sample()], rows)
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.metadata["RH_percent"] == [30.0, 45.0, 60.0]

    def test_the_raw_trace_paths_travel_with_the_numbers(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.metadata["raw_trace_file"] == [
            "traces/M-1.csv",
            "traces/M-2.csv",
            "traces/M-3.csv",
        ]

    def test_the_instrument_is_taken_from_the_measurement_rows(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series())
        (result,) = IteWorkbookParser().parse_all(path)
        assert result.instrument == "Keithley 6517B"

    def test_a_hand_typed_date_is_left_out_rather_than_guessed(self):
        """A silently mis-read date is worse than a missing one."""
        typed = _series()
        for row in typed:
            row["datetime_start"] = "19/08/2026 10:00"
        assert _first_datetime(typed, "datetime_start") == (None, False)

    def test_excel_timestamps_are_naive_and_get_utc_attached(self, tmp_path):
        """Excel cannot store a timezone, and ParsedData requires one.

        openpyxl refuses to write a tz-aware datetime at all, so every real
        workbook yields naive wall-clock. Dropping it would lose the ordering
        that makes drift across a campaign visible, so UTC is attached and the
        assumption is recorded once per sample rather than left implicit.
        """
        rows = _series()
        naive = datetime(2026, 9, 15, 10, 30)
        for row in rows:
            row["datetime_start"] = naive
        (result,) = IteWorkbookParser().parse_all(_write(tmp_path, [_sample()], rows))

        assert result.measured_at == naive.replace(tzinfo=UTC)
        notes = [i for i in result.issues if i.field == "datetime_start"]
        assert notes and notes[0].severity is Severity.INFO
        assert "no timezone" in notes[0].message


class TestMalformedFiles:
    def test_a_missing_sheet_is_an_error_not_a_crash(self, tmp_path):
        path = _write(tmp_path, [_sample()], _series(), sheets=("samples", "runs"))
        results = IteWorkbookParser().parse_all(path)
        assert len(results) == 1
        assert results[0].issues[0].severity is Severity.ERROR
        assert "measurements" in results[0].issues[0].message

    def test_an_empty_samples_sheet_is_an_error(self, tmp_path):
        path = _write(tmp_path, [], _series())
        results = IteWorkbookParser().parse_all(path)
        assert "no data rows" in results[0].issues[0].message

    def test_an_unopenable_file_never_raises(self, tmp_path):
        path = tmp_path / "broken.xlsx"
        path.write_bytes(b"not a workbook at all")
        results = IteWorkbookParser().parse_all(path)
        assert results[0].issues[0].severity is Severity.ERROR
        assert IteWorkbookParser().can_parse(path) == 0.0

    def test_metadata_is_json_safe(self, tmp_path):
        """ParsedData enforces this, so a datetime cell must be converted."""
        import json

        rows = _series()
        rows[0]["datetime_start"] = datetime(2026, 9, 15, 10, 30)
        path = _write(tmp_path, [_sample()], rows)
        (result,) = IteWorkbookParser().parse_all(path)
        json.dumps(result.metadata)  # must not raise


class TestTheGeneratorAndTheParserAgree:
    """The reason the schema and the writer live in one module.

    These two programs are the whole contract between the bench and the code.
    While the generator sat in an untracked scratch directory with its own copy
    of the column names, a rename on either side would have produced empty
    measurements rather than an error, and only a test written against the
    layout I happened to assume would have caught it.
    """

    def test_a_freshly_generated_template_is_recognised(self, tmp_path):
        path = write_template(tmp_path / "blank.xlsx")
        assert IteWorkbookParser().can_parse(path) == 1.0

    def test_a_blank_template_parses_to_no_samples_rather_than_a_crash(self, tmp_path):
        path = write_template(tmp_path / "blank.xlsx")
        results = IteWorkbookParser().parse_all(path)
        assert len(results) == 1
        assert "no data rows" in results[0].issues[0].message

    def test_every_column_the_parser_needs_exists_on_the_generated_sheet(self, tmp_path):
        """Guards the direction a rename would break silently."""
        import openpyxl

        from latos.ingestion.ite_workbook_template import (
            HEADER_ROW,
            MEASUREMENTS_SHEET,
            REQUIRED_MEASUREMENT_FIELDS,
            REQUIRED_SAMPLE_FIELDS,
            SAMPLES_SHEET,
        )

        wb = openpyxl.load_workbook(write_template(tmp_path / "blank.xlsx"))
        for sheet, required in (
            (SAMPLES_SHEET, REQUIRED_SAMPLE_FIELDS),
            (MEASUREMENTS_SHEET, REQUIRED_MEASUREMENT_FIELDS),
        ):
            header = {c.value for c in wb[sheet][HEADER_ROW]}
            assert set(required) <= header, f"{sheet} is missing {set(required) - header}"

    def test_the_test_fixtures_use_the_real_column_names(self):
        """Keeps these tests honest.

        Everything above builds its own workbook rather than the generated one,
        for speed and control. That only proves anything if the column names it
        invents are the real ones.
        """
        assert set(SAMPLE_COLS) <= {c.name for c in SAMPLE_COLUMNS}
        assert set(MEAS_COLS) <= {c.name for c in MEASUREMENT_COLUMNS}

    def test_required_fields_are_a_subset_of_the_tier_one_columns(self):
        """The parser may only insist on fields the template asks the bench for."""
        from latos.ingestion.ite_workbook_template import (
            REQUIRED_MEASUREMENT_FIELDS,
            REQUIRED_SAMPLE_FIELDS,
        )

        for columns, required in (
            (SAMPLE_COLUMNS, REQUIRED_SAMPLE_FIELDS),
            (MEASUREMENT_COLUMNS, REQUIRED_MEASUREMENT_FIELDS),
        ):
            tier1 = {c.name for c in columns if c.tier == 1}
            assert set(required) <= tier1
