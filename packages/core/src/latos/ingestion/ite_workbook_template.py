"""The ionic-thermoelectric recording workbook: its schema, and the writer for it.

The workbook is the contract between the bench and the code. Two programs have
to agree on it exactly — the one that writes the blank template and the one that
reads it back — and for a while they did not: the generator lived in an
untracked scratch directory while the parser carried its own copy of the sheet
names, the row offsets and every column heading. Nothing would have failed
loudly if the two drifted. A renamed column would simply have produced empty
measurements, and only a test written against the assumed layout would have
noticed.

So the columns are declared once, here, and both directions read that
declaration. `write_template()` builds the blank workbook from it;
`ingestion.parsers.ite_workbook` imports the same constants to read a filled one
back. A change to a column name is now a change in one place that both sides
inherit, and a test round-trips the generator's own output through the parser.

Two notions of "required" live side by side, and they are deliberately not the
same list:

* `tier` is what the **bench** must record. Tier 1 fields cannot be
  reconstructed once the session is over, so the template shades them and the
  field guide says why.
* `REQUIRED_*_FIELDS` is what the **parser** reports as missing. It is a subset.
  A seed sample legitimately has no pre-registered prediction, because there is
  no model yet to make one, so demanding `predicted_S_mV_K` on every row would
  flag the first two samples of every campaign for a fault they cannot avoid.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

__all__ = [
    "FIRST_DATA_ROW",
    "HEADER_ROW",
    "MEASUREMENTS_SHEET",
    "MEASUREMENT_COLUMNS",
    "REQUIRED_MEASUREMENT_FIELDS",
    "REQUIRED_SAMPLE_FIELDS",
    "SAMPLES_SHEET",
    "SAMPLE_COLUMNS",
    "Column",
    "write_template",
]

SAMPLES_SHEET = "samples"
MEASUREMENTS_SHEET = "measurements"
GUIDE_SHEET = "read me first"

# Row 1 is the sheet title, row 2 the header, row 3 the unit hints. Data starts
# at row 4. The parser keys off these, which is the whole reason they are here.
TITLE_ROW = 1
HEADER_ROW = 2
UNIT_ROW = 3
FIRST_DATA_ROW = 4

# Blank rows pre-styled so the sheet is usable the moment it is opened.
_BLANK_ROWS = 20

TIER_UNRECOVERABLE = 1  # gone once the session ends
TIER_PREPARATION = 2  # recoverable from the sample or the notebook
TIER_BOOKKEEPING = 3  # useful context, not load-bearing
TIER_DERIVED = 0  # computed, never typed


@dataclass(frozen=True, slots=True)
class Column:
    """One column, described once for both the writer and the reader."""

    name: str
    width: float
    tier: int
    unit: str
    why: str


SAMPLE_COLUMNS: tuple[Column, ...] = (
    Column("sample_id", 14, 1, "", "Unique. Suggested form IL-001, IL-002."),
    Column("date_prepared", 13, 1, "YYYY-MM-DD", ""),
    Column("IL_A_name", 18, 1, "", "Full name including the anion."),
    Column("IL_B_name", 18, 1, "", "Full name including the anion."),
    Column("mass_IL_A_mg", 13, 1, "mg", "Actual weighed mass, not the target."),
    Column("mass_IL_B_mg", 13, 1, "mg", "Actual weighed mass, not the target."),
    Column("mass_fraction_x", 14, 0, "-", "DERIVED: mass_A / (mass_A + mass_B)."),
    Column("molar_mass_A", 13, 2, "g/mol", "Record once; needed for the mixing law."),
    Column("molar_mass_B", 13, 2, "g/mol", "Record once; needed for the mixing law."),
    Column("mole_fraction_x", 14, 0, "-", "DERIVED. The physically meaningful axis."),
    Column("fabric_type", 15, 2, "", "Cotton grade / supplier."),
    Column("fabric_lot", 12, 2, "", "Changing lot mid-campaign is a confounder."),
    Column("fabric_areal_density", 15, 2, "g/m2", ""),
    Column("fabric_thickness_mm", 15, 2, "mm", ""),
    Column("fabric_length_mm", 14, 2, "mm", ""),
    Column("fabric_width_mm", 14, 2, "mm", ""),
    Column("mass_fabric_dry_mg", 15, 2, "mg", "Weigh before soaking."),
    Column("mass_fabric_soaked_mg", 17, 2, "mg", "Weigh after soaking and blotting."),
    Column("IL_loading_mg_cm2", 15, 0, "mg/cm2", "DERIVED. A hidden second variable if it drifts."),
    Column("soak_time_min", 12, 2, "min", ""),
    Column("blotting_method", 16, 2, "", "Be consistent. Free text, but use the same words."),
    Column("glass_slide_size_mm", 16, 2, "mm", ""),
    Column("predicted_S_mV_K", 15, 1, "mV/K", "PRE-REGISTER: prediction, written BEFORE mixing."),
    Column("predicted_S_sigma", 15, 1, "mV/K", "PRE-REGISTER: the stated uncertainty."),
    Column("prediction_timestamp", 18, 1, "", "PRE-REGISTER: when the prediction was recorded."),
    Column("notes", 34, 3, "", "Anything unusual. Colour, smell, wetting, spills."),
)

MEASUREMENT_COLUMNS: tuple[Column, ...] = (
    Column("meas_id", 12, 1, "", "Unique per row."),
    Column("sample_id", 13, 1, "", "Must match a row in the samples sheet."),
    Column("datetime_start", 17, 1, "YYYY-MM-DD hh:mm", "Excel stores no timezone; read as UTC."),
    Column("replicate_index", 14, 1, "", "1, 2, 3 ... for repeats of the same condition."),
    Column(
        "fresh_or_remeasure",
        17,
        1,
        "fresh / remeasure",
        "A re-measured sample has aged. It is not an independent replicate.",
    ),
    Column("order_in_session", 15, 1, "", "1, 2, 3 ... Used to detect drift within a day."),
    Column("RH_percent", 11, 1, "%", "Humidity changes both sigma and |S|."),
    Column("T_ambient_C", 12, 1, "degC", ""),
    Column("T_hot_C", 10, 1, "degC", "Measured, not the set-point."),
    Column("T_cold_C", 10, 1, "degC", "Measured, not the set-point."),
    Column("delta_T_K", 11, 0, "K", "DERIVED: T_hot - T_cold."),
    Column("wait_time_s", 12, 1, "s", "From applying dT to reading V."),
    Column(
        "steady_state_reached",
        18,
        1,
        "yes / no / unsure",
        "Judge from the V(t) trace, not from impatience.",
    ),
    Column("delta_V_mV", 12, 1, "mV", "Fix the polarity convention and never change it."),
    Column("polarity_convention", 18, 1, "", "e.g. 'V+ lead on cold electrode'. Same every run."),
    Column("electrode_material", 17, 1, "", "Sets the sign. Never change mid-campaign."),
    Column("electrode_spacing_mm", 18, 1, "mm", "Equilibration time scales as the square of this."),
    Column("voltmeter_model", 16, 1, "", ""),
    Column(
        "input_impedance_ohm",
        18,
        1,
        "ohm",
        "Below ~10 GOhm the meter loads the cell and reads low.",
    ),
    Column("sampling_interval_s", 17, 2, "s", ""),
    Column(
        "raw_trace_file",
        30,
        1,
        "",
        "Path to the V(t) and T(t) file. A single number cannot be audited.",
    ),
    Column("tau_fitted_s", 12, 0, "s", "DERIVED from the trace. Should vary smoothly with x."),
    Column("operator", 12, 3, "", ""),
    Column("flags", 26, 3, "", "Anything that went wrong. Write it down even if minor."),
)

# What the parser reports as missing. A strict subset of the Tier-1 columns:
# see the module docstring for why the pre-registration fields are not here.
REQUIRED_SAMPLE_FIELDS: tuple[str, ...] = ("mass_IL_A_mg", "mass_IL_B_mg")
REQUIRED_MEASUREMENT_FIELDS: tuple[str, ...] = (
    "RH_percent",
    "T_hot_C",
    "T_cold_C",
    "wait_time_s",
    "delta_V_mV",
    "electrode_material",
)

# Free-text columns whose values must stay consistent across a campaign, so the
# sheet offers a closed list rather than inviting "yes"/"Yes"/"y".
_VALIDATIONS = {
    "steady_state_reached": '"yes,no,unsure"',
    "fresh_or_remeasure": '"fresh,remeasure"',
}

_HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
_TIER1_FILL = PatternFill("solid", fgColor="FBE9E7")
_DERIVED_FILL = PatternFill("solid", fgColor="EEF3F8")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
_BODY_FONT = Font(size=10, name="Calibri")
_NOTE_FONT = Font(size=9, name="Calibri", italic=True, color="666666")
_THIN = Side(style="thin", color="D0D7DE")
_EDGE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_GUIDE = (
    ("Ionic thermoelectric recording template", True),
    ("", False),
    (
        "Fill this in at the bench, not afterwards from memory. Generated by "
        "latos.ingestion.ite_workbook_template, and read back by the ite-workbook parser.",
        False,
    ),
    ("", False),
    (
        "Pink columns cannot be reconstructed after the run. If one is blank, that run "
        "cannot be interpreted later and nothing recovers it.",
        False,
    ),
    ("Blue columns are derived. Leave them; Latos computes them from what you typed.", False),
    ("", False),
    ("THE THREE RULES", True),
    (
        "1. Measure every composition at three or more delta-T values. The Seebeck "
        "coefficient is the SLOPE of delta-V against delta-T, and the intercept tells you "
        "how much of the signal was electrode polarisation rather than thermoelectric. A "
        "single delta-T point cannot separate them.",
        False,
    ),
    (
        "2. Save the full V(t) trace every time. It is the only evidence that steady state "
        "was reached. Reported build-up times are 500 to 5000 seconds.",
        False,
    ),
    (
        "3. Write the predicted value and its uncertainty into the samples sheet BEFORE "
        "mixing. That is the pre-registration, and it cannot be added later.",
        False,
    ),
    ("", False),
    ("BEFORE THE CAMPAIGN", True),
    (
        "Measure both pure liquids first, x = 0 and x = 1. If their Seebeck coefficients "
        "have opposite signs, the mixture is worse than both endpoints and the target "
        "should change to the zero crossing. Decide that before spending cycles.",
        False,
    ),
    ("", False),
    ("FIELD REFERENCE", True),
)


def write_template(path: Path) -> Path:
    """Write the blank recording workbook to `path` and return it."""
    wb = Workbook()
    guide = wb.active
    guide.title = GUIDE_SHEET
    _build_guide(guide)

    for sheet_name, columns, title in (
        (SAMPLES_SHEET, SAMPLE_COLUMNS, "One row per prepared sample"),
        (
            MEASUREMENTS_SHEET,
            MEASUREMENT_COLUMNS,
            "One row per measurement point. Three or more per sample, one for each delta-T.",
        ),
    ):
        _build_sheet(wb.create_sheet(sheet_name), columns, title)

    wb.save(path)
    return path


def _build_sheet(ws: Any, columns: tuple[Column, ...], title: str) -> None:
    ws.cell(row=TITLE_ROW, column=1, value=title).font = Font(
        bold=True, size=11, color="1F4E79", name="Calibri"
    )
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=len(columns))
    ws.row_dimensions[TITLE_ROW].height = 20

    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        ws.column_dimensions[letter].width = column.width

        head = ws.cell(row=HEADER_ROW, column=index, value=column.name)
        head.font = _HEAD_FONT
        head.fill = _HEAD_FILL
        head.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        head.border = _EDGE

        unit = ws.cell(row=UNIT_ROW, column=index, value=column.unit)
        unit.font = _NOTE_FONT
        unit.alignment = Alignment(horizontal="center")
        unit.border = _EDGE

        fill = _fill_for(column.tier)
        if fill is not None:
            unit.fill = fill
        for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + _BLANK_ROWS):
            cell = ws.cell(row=row, column=index)
            cell.font = _BODY_FONT
            cell.border = _EDGE
            if fill is not None:
                cell.fill = fill

        if column.name in _VALIDATIONS:
            rule = DataValidation(type="list", formula1=_VALIDATIONS[column.name], allow_blank=True)
            ws.add_data_validation(rule)
            rule.add(f"{letter}{FIRST_DATA_ROW}:{letter}200")

    ws.row_dimensions[HEADER_ROW].height = 30
    ws.freeze_panes = f"A{FIRST_DATA_ROW}"


def _fill_for(tier: int) -> PatternFill | None:
    if tier == TIER_UNRECOVERABLE:
        return _TIER1_FILL
    if tier == TIER_DERIVED:
        return _DERIVED_FILL
    return None


def _build_guide(ws: Any) -> None:
    for letter, width in (("A", 26), ("B", 12), ("C", 16), ("D", 74)):
        ws.column_dimensions[letter].width = width

    row = 1
    for text, is_heading in _GUIDE:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(
            bold=is_heading,
            size=11 if is_heading else 10,
            color="1F4E79" if is_heading else "2B2B2B",
            name="Calibri",
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        if text and not is_heading:
            ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    labels = {1: "TIER 1", 2: "tier 2", 3: "tier 3", 0: "derived"}
    for sheet_name, columns in (
        (SAMPLES_SHEET, SAMPLE_COLUMNS),
        (MEASUREMENTS_SHEET, MEASUREMENT_COLUMNS),
    ):
        title = ws.cell(row=row, column=1, value=f"sheet: {sheet_name}")
        title.font = Font(bold=True, size=10, color="1F4E79", name="Calibri")
        row += 1
        for index, label in enumerate(("field", "tier", "unit", "why"), start=1):
            head = ws.cell(row=row, column=index, value=label)
            head.font = _HEAD_FONT
            head.fill = _HEAD_FILL
            head.border = _EDGE
        row += 1
        for column in columns:
            values = (column.name, labels[column.tier], column.unit, column.why)
            fill = _fill_for(column.tier)
            for index, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=index, value=value)
                cell.font = _BODY_FONT
                cell.border = _EDGE
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill is not None:
                    cell.fill = fill
            row += 1
        row += 1
