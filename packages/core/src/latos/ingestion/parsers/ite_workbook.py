"""Reader for the ionic-thermoelectric lab recording workbook.

File format
-----------
The workbook written by `docs/make_ite_template.py`, filled in at the bench.
Three sheets, of which two carry data:

* **samples** — one row per prepared sample: which two ionic liquids, the
  masses actually weighed, the fabric and its loading, and the prediction
  pre-registered before mixing.
* **measurements** — one row per (sample, ΔT) point: humidity, both measured
  temperatures, the voltage, the electrode, and a path to the raw trace.

The join between them is `sample_id`. That shape is deliberate: a sample is
prepared once and measured several times, and flattening the two into one sheet
would force every preparation field to be retyped per temperature point, which
is how transcription errors enter a dataset.

Why this is a parser rather than a bespoke importer
---------------------------------------------------
`BaseParser.parse_all` already returns one `ParsedData` per measurement, and the
orchestrator creates one `Measurement` from each. A workbook holding N samples
therefore maps cleanly onto N measurements, and the whole provenance chain —
file hash, parser version, re-parse caching — comes for free. The framework was
built for exactly this case; `thermoelectric_xlsx` uses it for one-sheet-per-
sample workbooks.

Each emitted `ParsedData` carries that sample's (ΔT, ΔV) series as arrays, which
is precisely what `analysis.thermovoltage.slope` consumes. Preparation fields
travel in `metadata` so they stay attached to the numbers they explain.

Validation policy
-----------------
Missing Tier-1 fields are the point of the exercise, not an edge case. A run
recorded without humidity or without a wait time cannot be interpreted later and
cannot be reconstructed, so each absence is reported as its own issue naming the
field and the row. Nothing is silently defaulted: inventing a plausible humidity
would produce a dataset that looks complete and is not.

A sample whose rows cannot yield even two usable (ΔT, ΔV) pairs is emitted with
its issues and no arrays, rather than dropped. A row that vanishes without
comment is indistinguishable from a row that was never written.
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any, ClassVar

import numpy as np
import openpyxl

from latos.core.enums import Severity, Technique
from latos.core.models import ValidationIssue, utc_now
from latos.ingestion.base_parser import BaseParser
from latos.ingestion.parsed_data import ParsedData

__all__ = ["IteWorkbookParser"]

_SAMPLES_SHEET = "samples"
_MEASUREMENTS_SHEET = "measurements"

# The header sits on row 2; row 1 is the sheet title and row 3 the unit hints.
_HEADER_ROW = 2
_FIRST_DATA_ROW = 4

# Fields that cannot be reconstructed after the run. Their absence is reported
# per row, because a campaign is only as analysable as its worst-recorded run.
_TIER1_MEASUREMENT = (
    "RH_percent",
    "T_hot_C",
    "T_cold_C",
    "wait_time_s",
    "delta_V_mV",
    "electrode_material",
)
_TIER1_SAMPLE = ("mass_IL_A_mg", "mass_IL_B_mg")

# Preparation columns copied into metadata so they stay with the numbers they
# explain. Everything else on the sheet is free text for a human reader.
_SAMPLE_METADATA = (
    "IL_A_name",
    "IL_B_name",
    "mass_IL_A_mg",
    "mass_IL_B_mg",
    "molar_mass_A",
    "molar_mass_B",
    "fabric_type",
    "fabric_lot",
    "fabric_thickness_mm",
    "mass_fabric_dry_mg",
    "mass_fabric_soaked_mg",
    "soak_time_min",
    "predicted_S_mV_K",
    "predicted_S_sigma",
    "prediction_timestamp",
    "notes",
)

# Per-measurement columns worth keeping alongside the arrays. Recorded as
# lists, one entry per ΔT point, so a humidity drift across a series stays
# visible rather than being averaged into a single number.
_MEASUREMENT_METADATA = (
    "RH_percent",
    "T_ambient_C",
    "wait_time_s",
    "steady_state_reached",
    "electrode_material",
    "electrode_spacing_mm",
    "input_impedance_ohm",
    "raw_trace_file",
    "replicate_index",
    "fresh_or_remeasure",
)

_MIN_POINTS = 2


class IteWorkbookParser(BaseParser):
    """One measurement per sample from the ionic-TE recording workbook."""

    name: ClassVar[str] = "ite-workbook"
    version: ClassVar[str] = "1.0.0"
    technique: ClassVar[Technique] = Technique.THERMOELECTRIC
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx",)

    def can_parse(self, path: Path) -> float:
        """Cheap check: both data sheets present, with the expected header row.

        Sheet names alone would be too eager — plenty of workbooks have a sheet
        called "samples". Confirming `sample_id` on the header row of each makes
        a false positive very unlikely without reading any data.
        """
        if not self._extension_matches(path):
            return 0.0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return 0.0
        try:
            names = set(wb.sheetnames)
            if not {_SAMPLES_SHEET, _MEASUREMENTS_SHEET} <= names:
                return 0.0
            for sheet in (_SAMPLES_SHEET, _MEASUREMENTS_SHEET):
                if "sample_id" not in _header(wb[sheet]):
                    return 0.0
        finally:
            wb.close()
        return 1.0

    def parse(self, path: Path) -> ParsedData:
        """First sample only; `parse_all` is the real entry point."""
        results = self.parse_all(path)
        return results[0] if results else self._empty(_open_failure("no samples found"))

    def parse_all(self, path: Path) -> tuple[ParsedData, ...]:
        """One `ParsedData` per sample row, carrying that sample's ΔT series."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return (self._empty(_open_failure(f"could not open workbook: {exc}")),)

        try:
            missing = [s for s in (_SAMPLES_SHEET, _MEASUREMENTS_SHEET) if s not in wb.sheetnames]
            if missing:
                return (
                    self._empty(
                        _open_failure(
                            f"workbook is missing the {', '.join(missing)} sheet(s); "
                            f"found {', '.join(wb.sheetnames)}"
                        )
                    ),
                )

            samples = _rows(wb[_SAMPLES_SHEET])
            measurements = _rows(wb[_MEASUREMENTS_SHEET])
        finally:
            wb.close()

        if not samples:
            return (self._empty(_open_failure("the samples sheet has no data rows")),)

        by_sample: dict[str, list[dict[str, Any]]] = {}
        orphans: list[dict[str, Any]] = []
        known = {str(row.get("sample_id")) for row in samples if row.get("sample_id") is not None}
        for row in measurements:
            sample_id = row.get("sample_id")
            if sample_id is None:
                continue
            key = str(sample_id)
            if key in known:
                by_sample.setdefault(key, []).append(row)
            else:
                orphans.append(row)

        return tuple(
            self._build(row, by_sample.get(str(row.get("sample_id")), []), orphans)
            for row in samples
            if row.get("sample_id") is not None
        )

    # ─── Internals ───────────────────────────────────────────────────
    def _build(
        self,
        sample: dict[str, Any],
        rows: list[dict[str, Any]],
        orphans: list[dict[str, Any]],
    ) -> ParsedData:
        sample_id = str(sample["sample_id"])
        issues: list[ValidationIssue] = []

        for field in _TIER1_SAMPLE:
            if _blank(sample.get(field)):
                issues.append(
                    _issue(
                        field,
                        f"Sample {sample_id}: {field} is empty. The mixing ratio derives "
                        "from the masses actually weighed, so it cannot be recovered.",
                    )
                )

        deltas, volts, issues_from_rows = _series(sample_id, rows)
        issues.extend(issues_from_rows)

        if orphans:
            ids = sorted({str(r.get("sample_id")) for r in orphans})
            issues.append(
                _issue(
                    "sample_id",
                    f"{len(orphans)} measurement row(s) reference sample id(s) not on the "
                    f"samples sheet: {', '.join(ids)}. Those rows were not attached to any "
                    "sample.",
                    severity=Severity.WARNING,
                )
            )

        arrays: dict[str, np.ndarray] = {}
        if deltas.size >= _MIN_POINTS:
            arrays = {"delta_t_k": deltas, "delta_v_mv": volts}
        else:
            issues.append(
                _issue(
                    "measurements",
                    f"Sample {sample_id}: only {deltas.size} usable (ΔT, ΔV) point(s). "
                    "The Seebeck coefficient is a fitted slope, so it needs at least "
                    f"{_MIN_POINTS}, and three or more to separate the electrode offset.",
                )
            )

        measured_at, naive = _first_datetime(rows, "datetime_start")
        if naive:
            issues.append(
                _issue(
                    "datetime_start",
                    f"Sample {sample_id}: the workbook's timestamps carry no timezone, "
                    "because Excel cannot store one. They are recorded as UTC, so the "
                    "ordering is right but the absolute time may be offset from the "
                    "wall clock at the bench.",
                    severity=Severity.INFO,
                )
            )

        metadata: dict[str, Any] = {"sample_id": sample_id}
        metadata.update(_scalars(sample, _SAMPLE_METADATA))
        metadata.update(_mass_fraction(sample))
        metadata.update(_per_point(rows, _MEASUREMENT_METADATA))

        return ParsedData(
            technique=self.technique,
            arrays=arrays,
            metadata=metadata,
            instrument=_first_str(rows, "voltmeter_model"),
            measured_at=measured_at,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
            features=_features(sample, metadata),
        )

    def _empty(self, issues: tuple[ValidationIssue, ...]) -> ParsedData:
        return ParsedData(
            technique=self.technique,
            arrays={},
            metadata={},
            instrument=None,
            measured_at=None,
            issues=issues,
            parser_name=self.name,
            parser_version=self.version,
        )


def _header(sheet: Any) -> dict[str, int]:
    """Column name → zero-based index, read from the header row."""
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == _HEADER_ROW:
            return {str(v).strip(): i for i, v in enumerate(row) if v is not None}
        if index > _HEADER_ROW:
            break
    return {}


def _rows(sheet: Any) -> list[dict[str, Any]]:
    """Every populated data row as a `{column: value}` dict."""
    header: dict[str, int] = {}
    out: list[dict[str, Any]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == _HEADER_ROW:
            header = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
            continue
        if index < _FIRST_DATA_ROW or not header:
            continue
        record = {name: row[i] if i < len(row) else None for name, i in header.items()}
        if any(not _blank(v) for v in record.values()):
            out.append(record)
    return out


def _series(
    sample_id: str, rows: list[dict[str, Any]]
) -> tuple[np.ndarray, np.ndarray, list[ValidationIssue]]:
    """Extract (ΔT, ΔV) pairs, reporting every row that could not supply one."""
    issues: list[ValidationIssue] = []
    deltas: list[float] = []
    volts: list[float] = []

    for row in rows:
        label = f"Sample {sample_id}, measurement {row.get('meas_id') or '(unlabelled)'}"
        for field in _TIER1_MEASUREMENT:
            if _blank(row.get(field)):
                issues.append(
                    _issue(field, f"{label}: {field} is empty and cannot be recovered later.")
                )

        hot, cold, volt = (_number(row.get(k)) for k in ("T_hot_C", "T_cold_C", "delta_V_mV"))
        if hot is None or cold is None or volt is None:
            continue
        # A difference in Celsius is a difference in kelvin; no offset applies.
        deltas.append(hot - cold)
        volts.append(volt)

    return (
        np.asarray(deltas, dtype=float),
        np.asarray(volts, dtype=float),
        issues,
    )


def _mass_fraction(sample: dict[str, Any]) -> dict[str, Any]:
    """Derive the composition knob from the masses actually weighed.

    Recomputed here rather than read from the sheet's own column, so a stale
    formula or a hand-typed value cannot disagree with the masses beside it.
    """
    a, b = _number(sample.get("mass_IL_A_mg")), _number(sample.get("mass_IL_B_mg"))
    if a is None or b is None or (a + b) <= 0:
        return {}
    return {"mass_fraction_x": round(a / (a + b), 6)}


def _features(sample: dict[str, Any], metadata: dict[str, Any]) -> dict[str, float]:
    """Curated scalars surfaced on the Measurement itself."""
    features: dict[str, float] = {}
    if "mass_fraction_x" in metadata:
        features["mass_fraction_x"] = float(metadata["mass_fraction_x"])
    for key in ("predicted_S_mV_K", "predicted_S_sigma"):
        value = _number(sample.get(key))
        if value is not None:
            features[key] = value
    return features


def _scalars(row: dict[str, Any], fields: tuple[str, ...]) -> dict[str, Any]:
    return {f: _json_safe(row.get(f)) for f in fields if not _blank(row.get(f))}


def _per_point(rows: list[dict[str, Any]], fields: tuple[str, ...]) -> dict[str, Any]:
    """One list per field, aligned with the ΔT series.

    Lists rather than a single representative value: humidity drifting across a
    series is exactly the confounder the campaign-level check looks for, and
    averaging it here would hide it.
    """
    out: dict[str, Any] = {}
    for f in fields:
        values = [_json_safe(r.get(f)) for r in rows]
        if any(v is not None for v in values):
            out[f] = values
    return out


def _number(value: Any) -> float | None:
    """A float, or None when the cell holds anything that is not one."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _json_safe(value: Any) -> Any:
    """Cell values reduced to the types `ParsedData.metadata` allows."""
    if value is None or isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _first_str(rows: list[dict[str, Any]], field: str) -> str | None:
    for row in rows:
        value = row.get(field)
        if not _blank(value):
            return str(value)
    return None


def _first_datetime(rows: list[dict[str, Any]], field: str) -> tuple[datetime | None, bool]:
    """First real timestamp in the series, and whether it arrived without a zone.

    Only a genuine datetime counts. A hand-typed string could be parsed with a
    guessed format, and a silently mis-read date is worse than a missing one.

    Excel cannot store a timezone — openpyxl refuses to write one — so every
    timestamp a filled workbook yields is naive wall-clock. `ParsedData` requires
    an aware value, which leaves three options: drop the timestamp, invent a
    zone, or attach one and say so. Dropping it loses the ordering that makes
    drift across a campaign visible, and ordering survives any consistent
    choice. So UTC is attached and the assumption is reported once per sample,
    because a lab in JST would otherwise find every measurement recorded nine
    hours from when it happened, with nothing on record to explain it.

    Returns `(timestamp, was_naive)`.
    """
    for row in rows:
        value = row.get(field)
        if isinstance(value, datetime):
            if value.tzinfo is None:
                return value.replace(tzinfo=UTC), True
            return value, False
    return None, False


def _issue(field: str, message: str, severity: Severity = Severity.ERROR) -> ValidationIssue:
    return ValidationIssue(field=field, severity=severity, message=message, detected_at=utc_now())


def _open_failure(message: str) -> tuple[ValidationIssue, ...]:
    return (_issue("file", message),)
