"""Resistivity + Seebeck parser for Linseis LSR `.xlsx` exports.

File format
-----------
Excel `.xlsx`, one sheet. A long instrument header (date, sample,
dimensions, per-channel sensor blocks) precedes the data. The data
block has an *empty* first column and three numeric columns:

    <blank> | Temperature(°C) | Resistivity(µΩ·m) | Seebeck(µV/K)

The header rows never match this ``[None, num, num, num]`` shape, so we
key on it directly rather than trying to locate a header row.

Units (confirmed with the researcher for this instrument):
- Temperature: °C  → converted to Kelvin (``temperature_k``).
- Resistivity: µΩ·m → kept as ``resistivity_uohm_m`` (the transport
  kernel converts to Ω·m, an explicit, provenanced step).
- Seebeck: µV/K → kept as ``seebeck_uv_k``.

Sample identity
---------------
Grouped in a technique folder (``Resistivity and Seebeck/CS R and S .xlsx``),
so the sample is in the filename. The ``R and S`` token is stripped
(``"CS R and S"`` → ``"CS"``) and emitted as ``metadata["sample_name"]``
so it lines up with the same sample's LFA file.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, ClassVar

import numpy as np
import openpyxl

from latos.core.enums import Severity, Technique
from latos.core.models import ValidationIssue, utc_now
from latos.ingestion.base_parser import BaseParser
from latos.ingestion.parsed_data import ParsedData

__all__ = ["ResistivitySeebeckXlsxParser"]

# Rows to scan when sniffing in `can_parse` (data starts well down the sheet).
_SNIFF_ROWS = 80
# Distinct data rows required for a confident match.
_MIN_DATA_ROWS = 3
# Columns a data row must have: blank, T, resistivity, seebeck.
_MIN_DATA_COLS = 4
# Celsius → Kelvin offset.
_C_TO_K = 273.15
# Technique tokens stripped from the filename to recover the sample name.
_RS_TOKEN_RE = re.compile(r"\br\s*(?:and|&)\s*s\b", re.IGNORECASE)


def _is_data_row(row: tuple[Any, ...]) -> bool:
    """True for the Linseis data shape: empty col A, then 3 numbers."""
    if len(row) < _MIN_DATA_COLS:
        return False
    if row[0] is not None:
        return False
    return all(isinstance(row[i], int | float) for i in (1, 2, 3))


def _sample_name_from_path(path: Path) -> str:
    """``"CS R and S"`` → ``"CS"`` — strip the technique token from the stem."""
    cleaned = _RS_TOKEN_RE.sub(" ", path.stem)
    return " ".join(cleaned.split()).strip() or path.stem


class ResistivitySeebeckXlsxParser(BaseParser):
    """Parser for Linseis LSR resistivity + Seebeck `.xlsx` exports."""

    name: ClassVar[str] = "resistivity-seebeck-xlsx"
    version: ClassVar[str] = "1.0.0"
    technique: ClassVar[Technique] = Technique.THERMOELECTRIC
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx",)

    def can_parse(self, path: Path) -> float:
        """1.0 when the sheet mentions Seebeck and has ≥3 Linseis data rows."""
        if not self._extension_matches(path):
            return 0.0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return 0.0
        try:
            sheet = wb[wb.sheetnames[0]]
            data_rows = 0
            saw_seebeck = False
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= _SNIFF_ROWS:
                    break
                if any(isinstance(c, str) and "seebeck" in c.lower() for c in row):
                    saw_seebeck = True
                if _is_data_row(row):
                    data_rows += 1
            return 1.0 if saw_seebeck and data_rows >= _MIN_DATA_ROWS else 0.0
        finally:
            wb.close()

    def parse(self, path: Path) -> ParsedData:
        """Parse the resistivity + Seebeck data block into aligned arrays."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return self._empty([
                ValidationIssue(
                    field="file",
                    severity=Severity.ERROR,
                    message=f"Could not open workbook: {exc}",
                    detected_at=utc_now(),
                ),
            ])
        try:
            return self._parse_sheet(wb[wb.sheetnames[0]], path)
        finally:
            wb.close()

    # ─── Internals ───────────────────────────────────────────────────
    def _parse_sheet(self, sheet: Any, path: Path) -> ParsedData:
        temperature_k: list[float] = []
        resistivity: list[float] = []
        seebeck: list[float] = []
        for row in sheet.iter_rows(values_only=True):
            if not _is_data_row(row):
                continue
            temperature_k.append(float(row[1]) + _C_TO_K)
            resistivity.append(float(row[2]))
            seebeck.append(float(row[3]))

        if not temperature_k:
            return self._empty([
                ValidationIssue(
                    field="data",
                    severity=Severity.ERROR,
                    message="No resistivity/Seebeck data rows found.",
                    detected_at=utc_now(),
                ),
            ])

        arrays = {
            "temperature_k": np.asarray(temperature_k, dtype=np.float64),
            "resistivity_uohm_m": np.asarray(resistivity, dtype=np.float64),
            "seebeck_uv_k": np.asarray(seebeck, dtype=np.float64),
        }
        metadata: dict[str, Any] = {
            "sample_name": _sample_name_from_path(path),
            "measurement_kind": "resistivity_seebeck",
            "units": {
                "temperature_k": "K (converted from C)",
                "resistivity_uohm_m": "uOhm*m",
                "seebeck_uv_k": "uV/K",
            },
            "n_points": len(temperature_k),
            "temperature_k_min": min(temperature_k),
            "temperature_k_max": max(temperature_k),
        }
        return ParsedData(
            technique=self.technique,
            arrays=arrays,
            metadata=metadata,
            instrument="Linseis LSR (R&S xlsx export)",
            measured_at=None,
            issues=(),
            parser_name=self.name,
            parser_version=self.version,
        )

    def _empty(self, issues: list[ValidationIssue]) -> ParsedData:
        return ParsedData(
            technique=self.technique,
            arrays={},
            metadata={},
            instrument="Linseis LSR (R&S xlsx export)",
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )
