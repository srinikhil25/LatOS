"""LFA (Laser Flash Analysis) parser for `.xlsx` exports.

File format
-----------
Excel `.xlsx` with a single sheet. A few blank leading rows, then a
header row whose cells are prefixed with ``#``:

    #Temperature/K | #Model | #Diffusivity/(mm^2/s) | #Conductivity/(W/(m*K)) | #Cp-Calc/(J/(g*K))

followed by one numeric row per temperature. The ``#Model`` column is a
text label and is skipped.

Sample identity
---------------
These files live in a *technique* folder (``LFA/CS LFA.xlsx``), so the
sample name is in the filename, not the parent folder. The parser strips
the ``LFA`` token from the stem (``"CS LFA"`` → ``"CS"``) and emits it as
``metadata["sample_name"]`` so the orchestrator names the sample
correctly and the LFA + Resistivity/Seebeck files of one sample line up.

Validation policy: see `xrd_rigaku_txt.py` — same contract.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, ClassVar

import numpy as np
import openpyxl

from latos.core.enums import Severity, Technique
from latos.core.models import ValidationIssue, utc_now
from latos.ingestion.base_parser import BaseParser
from latos.ingestion.parsed_data import ParsedData

__all__ = ["LfaXlsxParser"]

# Rows to scan when sniffing for the header in `can_parse`.
_SNIFF_ROWS = 12
# Columns a data row must have: T(K), model, diffusivity, conductivity.
_MIN_DATA_COLS = 4
# The LFA token to strip from the filename to recover the sample name.
_LFA_TOKEN_RE = re.compile(r"\blfa\b", re.IGNORECASE)


def _norm(cell: Any) -> str:
    """Lowercased, ``#``-stripped text of a header cell (``""`` if not text)."""
    return cell.lstrip("#").strip().lower() if isinstance(cell, str) else ""


def _is_header_row(row: tuple[Any, ...]) -> bool:
    """True if this row carries the LFA diffusivity + conductivity headers."""
    text = " | ".join(_norm(c) for c in row)
    return "diffusivity" in text and "conductivity" in text


def _sample_name_from_path(path: Path) -> str:
    """``"CS LFA"`` → ``"CS"`` — strip the technique token from the stem."""
    cleaned = _LFA_TOKEN_RE.sub(" ", path.stem)
    return " ".join(cleaned.split()).strip() or path.stem


class LfaXlsxParser(BaseParser):
    """Parser for LFA thermal-conductivity `.xlsx` exports."""

    name: ClassVar[str] = "lfa-xlsx"
    version: ClassVar[str] = "1.0.0"
    technique: ClassVar[Technique] = Technique.THERMOELECTRIC
    supported_extensions: ClassVar[tuple[str, ...]] = (".xlsx",)

    def can_parse(self, path: Path) -> float:
        """1.0 when a sheet has the distinctive diffusivity+conductivity header."""
        if not self._extension_matches(path):
            return 0.0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return 0.0
        try:
            sheet = wb[wb.sheetnames[0]]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= _SNIFF_ROWS:
                    break
                if _is_header_row(row):
                    return 1.0
            return 0.0
        finally:
            wb.close()

    def parse(self, path: Path) -> ParsedData:
        """Parse the single LFA sheet into temperature + κ + diffusivity + Cp."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            return self._empty(
                [
                    ValidationIssue(
                        field="file",
                        severity=Severity.ERROR,
                        message=f"Could not open workbook: {exc}",
                        detected_at=utc_now(),
                    ),
                ]
            )
        try:
            return self._parse_sheet(wb[wb.sheetnames[0]], path)
        finally:
            wb.close()

    # ─── Internals ───────────────────────────────────────────────────
    def _parse_sheet(self, sheet: Any, path: Path) -> ParsedData:
        col: dict[str, list[float]] = {
            "temperature_k": [],
            "diffusivity_mm2_s": [],
            "thermal_conductivity": [],
            "cp_j_gk": [],
        }
        seen_header = False
        for row in sheet.iter_rows(values_only=True):
            if not seen_header:
                if _is_header_row(row):
                    seen_header = True
                continue
            # data row: col 0 = T(K), col 1 = model text, 2 = diffusivity,
            # 3 = conductivity, 4 = Cp.
            if len(row) < _MIN_DATA_COLS or not isinstance(row[0], int | float):
                continue
            if not isinstance(row[3], int | float):
                continue
            col["temperature_k"].append(float(row[0]))
            col["diffusivity_mm2_s"].append(_num(row[2]))
            col["thermal_conductivity"].append(float(row[3]))
            col["cp_j_gk"].append(_num(row[4] if len(row) > _MIN_DATA_COLS else None))

        issues: list[ValidationIssue] = []
        if not col["temperature_k"]:
            issues.append(
                ValidationIssue(
                    field="data",
                    severity=Severity.ERROR,
                    message="No LFA data rows found below the header.",
                    detected_at=utc_now(),
                ),
            )
            return self._empty(issues)

        arrays = {k: np.asarray(v, dtype=np.float64) for k, v in col.items()}
        metadata: dict[str, Any] = {
            "sample_name": _sample_name_from_path(path),
            "measurement_kind": "lfa",
            "n_points": len(col["temperature_k"]),
            "temperature_k_min": min(col["temperature_k"]),
            "temperature_k_max": max(col["temperature_k"]),
        }
        return ParsedData(
            technique=self.technique,
            arrays=arrays,
            metadata=metadata,
            instrument="LFA (xlsx export)",
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )

    def _empty(self, issues: list[ValidationIssue]) -> ParsedData:
        return ParsedData(
            technique=self.technique,
            arrays={},
            metadata={},
            instrument="LFA (xlsx export)",
            measured_at=None,
            issues=tuple(issues),
            parser_name=self.name,
            parser_version=self.version,
        )


def _num(v: Any) -> float:
    """Coerce a cell to float, or NaN when missing/non-numeric."""
    return float(v) if isinstance(v, int | float) else float("nan")
